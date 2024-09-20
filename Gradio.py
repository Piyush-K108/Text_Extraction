import shutil
from docx import Document
import fitz  # PyMuPDF
import openpyxl
import json
from run_assistance import json_to_csv, process_files, save_output_to_txt, process_image_file
from history_openai import process_code_files
from asking_openai import process_questions_files
import gradio as gr
import email
from email import policy
import os
import csv
import pandas as pd
import json
import numpy as np
import zipfile


def table_to_json(input_file):
    try:
        print(input_file)

        # Check file extension and read accordingly
        if input_file.endswith('.csv'):
            df = pd.read_csv(input_file)
        elif input_file.endswith('.xlsx') or input_file.endswith('.xls'):
            df = pd.read_excel(input_file)
        else:
            return "Unsupported file format. Please use a CSV or Excel file."

        # Replace NaN with None
        df = df.replace({np.nan: None})

        # Convert DataFrame to JSON format
        json_data = df.to_dict(orient='records')

        # Generate a unique name for the output JSON file based on the input file name
        base_name = os.path.splitext(os.path.basename(input_file))[0]
        json_file = f'{base_name}_output.json'

        # Write JSON data to a file
        with open(json_file, 'w') as f:
            json.dump(json_data, f, indent=4)

        print(f"JSON file created: {json_file}")
        return json_data

    except Exception as e:
        return f"Error occurred: {str(e)}"


def csv_to_txt(csv_file, txt_file):
    """Convert a CSV file to a TXT file."""
    try:
        with open(csv_file, 'r', encoding='utf-8') as csvf, open(txt_file, 'w', encoding='utf-8') as txtf:
            reader = csv.reader(csvf)
            for row in reader:
                # Write the row values separated by tabs or spaces
                txtf.write("\t".join(row) + "\n")
        print(f"Converted {csv_file} to {txt_file}")
    except Exception as e:
        print(f"Error converting CSV to TXT: {str(e)}")


def excel_to_txt(excel_file, txt_file):
    """Convert an Excel file to a TXT file without using pandas."""
    try:
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(excel_file)

        # Open the output TXT file for writing
        with open(txt_file, 'w', encoding='utf-8') as txtf:
            # Iterate over each sheet in the workbook
            for sheet_name in workbook.sheetnames:
                # Get the current sheet
                sheet = workbook[sheet_name]

                # Write the sheet name to the text file
                txtf.write(f"Sheet: {sheet_name}\n")

                # Write the header (column names)
                # Assumes first row is the header
                headers = [cell.value for cell in sheet[1]]
                txtf.write("\t".join([str(header)
                           for header in headers]) + "\n")

                # Write each row of data
                # Skip the header row
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_data = [
                        str(cell) if cell is not None else '' for cell in row]
                    txtf.write("\t".join(row_data) + "\n")

                # Add spacing between sheets
                txtf.write("\n\n")

        print(f"Converted Excel {excel_file} to {txt_file}")

    except Exception as e:
        print(f"Error converting Excel to TXT: {str(e)}")


def get_attachments_from_eml(basename, eml_file):
    # Directory where attachments will be saved
    output_dir = 'attachments/'

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    attachment_paths = []

    with open(eml_file, 'rb') as f:
        msg = email.message_from_binary_file(f, policy=policy.default)

    for part in msg.walk():
        # Check if the part is an attachment
        if part.get_content_disposition() == 'attachment':
            filename = part.get_filename()
            if filename:
                # Prefix the basename to the filename to avoid naming conflicts
                name, ext = os.path.splitext(filename)
                ext = ext.lower()

                full_filename = f"{basename}_{name}{ext}"
                attachment_path = os.path.join(output_dir, full_filename)

                # Save the attachment
                with open(attachment_path, 'wb') as fp:
                    fp.write(part.get_payload(decode=True))

                if ext == '.xls' or ext == '.xlsx':
                    txt_filename = os.path.splitext(full_filename)[0] + '.txt'
                    txt_path = os.path.join(output_dir, txt_filename)

                    if ext in ['.xls', '.xlsx']:
                        excel_to_txt(output_dir+"/"+full_filename, txt_path)

                    attachment_paths.append(txt_path)
                    print(
                        f"Saved and converted {ext.upper()} to TXT: {txt_path}")
                else:

                    attachment_paths.append(attachment_path)
                    print(f"Saved attachment: {attachment_path}")

    # Return the list of attachment paths
    return attachment_paths


def get_text_from_eml(basename, eml):
    output_dir = 'data/email_data/'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    with open(eml, 'rb') as f:
        msg = email.message_from_binary_file(f, policy=policy.default)

    email_body = ""
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == 'text/plain':
                email_body = part.get_payload(decode=True).decode(
                    part.get_content_charset())
                break
    else:
        email_body = msg.get_payload(decode=True).decode(
            msg.get_content_charset())

    text_file_path = os.path.join(output_dir, f'{basename}_text.txt')

    with open(text_file_path, 'w') as f:
        f.write(email_body)

    return text_file_path




def extract_and_remove_images_from_pdf(pdf_path):
    output_dir = 'images_in_att'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    output_file_dir = 'edited_att_without_images'
    if not os.path.exists(output_file_dir):
        os.makedirs(output_file_dir)

    image_paths = []
    output_pdf_path = os.path.join(output_file_dir, os.path.basename(pdf_path))

    try:
        doc = fitz.open(pdf_path)  # Open the PDF file

        # Iterate through all pages and extract images
        for page_num in range(doc.page_count):
            page = doc.load_page(page_num)
            images = page.get_images(full=True)

            for img_index, img in enumerate(images):
                xref = img[0]
                try:
                    # Extract the image bytes
                    base_image = doc.extract_image(xref)
                    image_bytes = base_image["image"]
                    image_ext = base_image["ext"]

                    # Save the image to the output folder
                    image_filename = f"image_{page_num + 1}_{img_index + 1}.{image_ext}" if image_ext else f"image_{page_num + 1}_{img_index + 1}.bin"
                    image_path = os.path.join(output_dir, image_filename)
                    with open(image_path, "wb") as image_file:
                        image_file.write(image_bytes)
                    image_paths.append(image_path)

                except Exception as img_err:
                    print(f"Error extracting image {img_index + 1} on page {page_num + 1}: {str(img_err)}")
                    continue  # Skip to the next image

        # Now remove images from the PDF
        for page_num in range(doc.page_count):
            page = doc.load_page(page_num)
            img_list = page.get_images(full=True)
            for img in img_list:
                try:
                    page.delete_image(img[0])  # Remove the image by its xref
                except Exception as e:
                    print(f"Error removing image on page {page_num + 1}: {str(e)}")

        # Save the modified PDF without images
        doc.save(output_pdf_path)
        doc.close()

    except Exception as e:
        print(f"Error processing PDF '{pdf_path}': {str(e)}")
        return False, None  # Return False if an error occurred

    if image_paths:
        return image_paths, output_pdf_path  # Return paths of extracted images and new PDF path
    else:
        return False, None  # Return False if no images were found



def extract_and_remove_images_from_docx(docx_path):
    output_dir = 'images_in_att'

    output_file_dir = 'edited_att_without_images'
    if not os.path.exists(output_file_dir):
        os.makedirs(output_file_dir)

    image_paths = []
    output_docx_path = os.path.join(output_file_dir, os.path.basename(docx_path))

    # Create output directory if it doesn't exist
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Temporary directory to unzip the DOCX content
    temp_dir = 'temp_docx'
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)
    os.makedirs(temp_dir)

    # Extract the .docx (which is essentially a zip archive) into a temporary folder
    with zipfile.ZipFile(docx_path, 'r') as docx_zip:
        docx_zip.extractall(temp_dir)

    image_paths = []
    media_dir = os.path.join(temp_dir, 'word', 'media')

    # Check if the media directory exists (it holds the images)
    if os.path.exists(media_dir):
        # Iterate over image files in the media directory
        for image_file in os.listdir(media_dir):
            image_path = os.path.join(media_dir, image_file)
            output_image_path = os.path.join(output_dir, image_file)

            # Move the image to the output directory
            shutil.move(image_path, output_image_path)
            image_paths.append(output_image_path)

        # Remove the media directory (since images are deleted)
        shutil.rmtree(media_dir)

    # Recreate the .docx without images by zipping the content
    with zipfile.ZipFile(output_docx_path, 'w') as new_docx_zip:
        for foldername, subfolders, filenames in os.walk(temp_dir):
            for filename in filenames:
                file_path = os.path.join(foldername, filename)
                archive_name = os.path.relpath(file_path, temp_dir)
                new_docx_zip.write(file_path, arcname=archive_name)

    # Clean up the temporary directory
    shutil.rmtree(temp_dir)

    if image_paths:
        # Return paths of extracted images and new DOCX path
        return image_paths, output_docx_path
    else:
        return False, None  # Return False if no images were found


def process_csv_excel(files):
    result_json = []

    try:

        for file in files:

            result = process_code_files(file.name)

            if result is not None:
                result_json.append(result)

    except Exception as e:
        print(f"Error in gradio_interface: {str(e)}")

    return result_json


def process_questions(files, txt):
    result_json = []
    try:
        for file in files:
            result = process_questions_files(file.name, txt)
            if result is not None:
                result_json.append(result)

    except Exception as e:
        print(f"Error in gradio_interface: {str(e)}")

    return result_json






def process_eml(file, basename):
    
    eml_result = []
    eml_text_path = get_text_from_eml(basename, file.name)
    eml_files = get_attachments_from_eml(basename, file.name)

    if eml_files:
        for attachment in eml_files:

            result2 = process_attachment(attachment)
            eml_result.append(result2)

    eml_result.append(process_files([eml_text_path]))
    return eml_result


def process_attachment(attachment):
    att_extension = os.path.basename(attachment).split('.')[-1]
    
    if att_extension == 'csv':
        return table_to_json(attachment)
    elif att_extension in ['png', 'jpeg', 'jpg']:
        txt_path_images = process_image_file(attachment)
        return process_files([txt_path_images])

    elif att_extension in ['pdf', 'docx']:
        
        return process_image_or_files(attachment, att_extension)
    
    else:
        return process_files([attachment])





def process_image_or_files(file_name, file_type):
    images_result = []
    if file_type == 'pdf':
        image_paths, pdf_path = extract_and_remove_images_from_pdf(file_name)
        for image_path in image_paths:
            images_txt_path = process_image_file(image_path) 
            images_result.append(process_files([images_txt_path]))
        images_result.append(process_files([pdf_path]))
    else:
        image_paths, doc_path = extract_and_remove_images_from_docx(file_name)
        for image_path in image_paths:
            images_txt_path = process_image_file(image_path) 
            images_result.append(process_files([images_txt_path]))
        images_result.append(process_files([doc_path]))

    return images_result


def process_file_by_extension(file, extension, basename):
    if extension == 'eml':
        
        return process_eml(file, basename)
    elif extension == 'csv':
        return table_to_json(file.name)
    elif extension in ['xls', 'xlsx']:
        output_dir = 'attachments/'
        os.makedirs(output_dir, exist_ok=True)
        txt_filename = basename + '.txt'
        txt_path = os.path.join(output_dir, txt_filename)
        excel_to_txt(file.name, txt_path)
        
        return process_files([txt_path])
    elif extension in ['png', 'jpg', 'jpeg']:
        txt_path_images = process_image_file(file.name)
        return process_files([txt_path_images])
    elif extension in ['pdf', 'docx']:
        return process_image_or_files(file.name, extension)
    else:
        return process_files([file.name])


def gradio_interface(files):
    result_json = []
    
    try:
        for file in files:
            file_name = os.path.basename(file)
            extension = file_name.split('.')[-1]
            basename = os.path.splitext(file_name)[0]

            result = process_file_by_extension(file, extension, basename)

            if result is not None:
                result_json.append(result)

    except Exception as e:
        print(f"Error in gradio_interface: {str(e)}")

    return result_json


csv_excel_interface = gr.Interface(
    fn=process_csv_excel,
    inputs=[gr.File(label="Upload CSV/XLSX Files", file_count="multiple")],
    outputs=gr.JSON(label="Extracted JSON Data (CSV/XLSX Files)"),
    description="Upload CSV or XLSX files to extract structured data for a specific date."
)


transcribing_interface = gr.Interface(
    fn=process_questions,
    inputs=[gr.File(label="Upload CSV/XLSX Files", file_count="multiple"), gr.Textbox(
        label="Enter your question", placeholder="Ask a question about the data")],
    outputs=gr.JSON(label="Extracted JSON Data (CSV/XLSX Files)"),
    description="Upload CSV or XLSX files to extract structured data for a specific date."
)


general_file_interface = gr.Interface(
    fn=gradio_interface,
    inputs=gr.File(label="Upload Files (Any Type)", file_count="multiple"),
    outputs=gr.JSON(label="Extracted JSON Data (General Files)"),
    description="Upload multiple files (any type) to extract structured data."
)




# Use TabbedInterface to group the two interfaces into tabs
app = gr.TabbedInterface(
    interface_list=[general_file_interface, transcribing_interface],
    tab_names=["General File Processing", "Question Processing"]
)

if __name__ == "__main__":
    app.launch()




























# def gradio_interface(files):
#     result_json = []
#     eml_result = []
#     images_result_in_att = []
#     try:

#         for file in files:

#             extension = os.path.basename(file).split('.')[-1]
#             file_name = os.path.basename(file)
#             basename = os.path.splitext(file_name)[0]

#             if extension == 'eml':
#                 eml_text_path = get_text_from_eml(basename, file.name)
#                 eml_files = get_attachments_from_eml(basename, file.name)

#                 if eml_files:
#                     print("files")
#                     for attachment in eml_files:
#                         att_extension = os.path.basename(
#                             attachment).split('.')[-1]
#                         att_basename = os.path.splitext(
#                             os.path.basename(attachment))[0]
#                         print(att_basename, att_extension)

#                         # Process each attachment based on its extension

#                         if att_extension == 'csv':
#                             table_to_json(attachment)
#                             result2 = 1

#                         elif att_extension == 'png' or att_extension == 'jpeg':
#                             txt_path_images = process_image_file(attachment)
#                             result2 = process_files([txt_path_images])

#                         else:
#                             result2 = process_files(eml_files)

#                         eml_result.append(result2)

#                 result2 = process_files([eml_text_path])
#                 eml_result.append(result2)
#                 result = eml_result

#             elif extension == 'csv':
#                 result = table_to_json(file.name)

#             elif extension in ['xls', 'xlsx']:

#                 output_dir = 'attachments/'

#                 if not os.path.exists(output_dir):
#                     os.makedirs(output_dir)

#                 txt_filename = basename + '.txt'
#                 txt_path = os.path.join(output_dir, txt_filename)
#                 excel_to_txt(file.name, txt_path)
#                 result = process_files([txt_path])

#             elif extension == 'png' or extension == 'jpg' or extension == 'jpeg':

#                 txt_path_images = process_image_file(file.name)
#                 result = process_files([txt_path_images])

#             elif extension == 'pdf':
#                 image_paths, pdf_path = extract_and_remove_images_from_pdf(file.name)

#                 for image_path in image_paths:
#                     result2 = process_image_file(image_path)
#                     images_result_in_att.append(result2)

#                 result2 = process_files([pdf_path])
#                 images_result_in_att.append(result2)
               
#                 result=images_result_in_att

#             elif extension == 'docx':
#                 image_path, doc_path = extract_and_remove_images_from_docx(file.name)
#                 for image_path in image_paths:
#                     result2 = process_image_file(image_path)
#                     images_result_in_att.append(result2)

#                 result2 = process_files([doc_path])
#                 images_result_in_att.append(result2)
               
#                 result=images_result_in_att
                
#             else:

#                 result = process_files([file.name])

#             if result is not None:
#                 result_json.append(result)

#     except Exception as e:
#         print(f"Error in gradio_interface: {str(e)}")

#     return result_json


